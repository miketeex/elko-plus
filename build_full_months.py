#!/usr/bin/env python3
"""
Byggir full-months.json: heilar mánaðartöflur (deildir, starfsmenn, vaktir,
samtölur) fyrir ELKO Vaktaplan, með feik nöfnum/símanúmerum.
Nær yfir alla mánuði frá upphafi og upp að (en ekki með) fyrsta "lifandi" mánuði
sem er þegar innfelldur í HTML-skránni (Júlí 2026).
"""
import sys, json, datetime
import openpyxl

MONTH_ORDER = ['Janúar','Febrúar','Mars','Apríl','Maí','Júní','Júlí','Ágúst','September','Október','Nóvember','Desember']
JS_WEEKDAYS = ['Sun','Mán','Þri','Mið','Fim','Fös','Lau']
CUTOFF = (2026, 7)  # sleppa mánuðum >= þessu (þeir eru þegar "lifandi" í HTML-inu)

def parse_month_sheet(name):
    for m in MONTH_ORDER:
        if name.startswith(m):
            try:
                year = int(name.split()[-1])
            except ValueError:
                return None, None
            return m, year
    return None, None

def find_label_row(ws, label, min_row=1, max_row=12, max_col=10):
    for r in range(min_row, max_row + 1):
        for c in range(1, max_col + 1):
            if ws.cell(row=r, column=c).value == label:
                return r, c
    return None, None

def status_kind(v):
    if v is None or str(v).strip() == '':
        return 'empty'
    s = str(v).strip().lower()
    if s == 'x':
        return 'unpaid'
    if s in ('frí', 'fri'):
        return 'summer'
    if 'f.orlof' in s:
        return 'parental'
    if s == 'v' or 'veik' in s:
        return 'sick'
    if any(ch.isdigit() for ch in s):
        return 'shift'
    return 'other'

def build_staff_map(wb):
    ws = wb['Staff']
    m = {}
    for r in range(2, ws.max_row + 1):
        sst = ws.cell(row=r, column=1).value
        dept = ws.cell(row=r, column=6).value
        if sst:
            m[str(sst).strip()] = str(dept).strip() if dept else ''
    return m

def build_full_months(xlsx_path, name_map, phone_map):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    staff_map = build_staff_map(wb)
    staff_initials = set(staff_map.keys())

    sheets = []
    for s in wb.sheetnames:
        clean = s.replace('Í vinnslu - ', '')
        m, y = parse_month_sheet(clean)
        if m:
            sheets.append((s, m, y, MONTH_ORDER.index(m)))
    sheets.sort(key=lambda x: (x[1], x[2]))

    months = {}
    skipped = []

    for sheetname, mname, y, midx in sheets:
        if (y, midx + 1) >= CUTOFF:
            continue  # þegar innfellt í HTML sem "lifandi" mánuður
        ws = wb[sheetname]
        row_v, label_col = find_label_row(ws, 'Vaktstjóri')
        if not row_v:
            skipped.append(sheetname)
            continue
        row_dags = None
        for rr in range(row_v + 1, row_v + 4):
            if ws.cell(row=rr, column=label_col).value == 'Dags':
                row_dags = rr
                break
        if not row_dags:
            skipped.append(sheetname)
            continue

        row_week, _ = find_label_row(ws, 'Vikunúmer', min_row=1, max_row=row_v, max_col=label_col)
        row_store, _ = find_label_row(ws, 'Verslanir', min_row=row_dags, max_row=row_dags + 5, max_col=label_col)
        row_back, _ = find_label_row(ws, 'Bakkinn', min_row=row_dags, max_row=row_dags + 6, max_col=label_col)
        row_campaign, _ = find_label_row(ws, 'Campaign', min_row=row_dags, max_row=row_dags + 8, max_col=label_col)
        row_event, _ = find_label_row(ws, 'Viðburður', min_row=row_dags, max_row=row_dags + 9, max_col=label_col)

        col_start = label_col + 1
        max_col = ws.max_column

        # dagasúlur
        day_cols = []
        col = col_start
        while col <= max_col:
            dv = ws.cell(row=row_dags, column=col).value
            if isinstance(dv, datetime.datetime) and dv.month == midx + 1 and dv.year == y:
                day_cols.append((col, dv))
            col += 1
        if not day_cols:
            skipped.append(sheetname)
            continue

        days = []
        for col, dv in day_cols:
            store_val = ws.cell(row=row_store, column=col).value if row_store else None
            store_str = str(store_val).strip() if store_val is not None else ''
            back_val = ws.cell(row=row_back, column=col).value if row_back else None
            back_str = str(back_val).strip() if back_val is not None else ''
            manager_val = ws.cell(row=row_v, column=col).value
            manager_str = str(manager_val).strip() if manager_val else ''
            week_val = ws.cell(row=row_week, column=col).value if row_week else ''
            campaign_val = ws.cell(row=row_campaign, column=col).value if row_campaign else ''
            event_val = ws.cell(row=row_event, column=col).value if row_event else ''
            closed = 'lokað' in store_str.lower() if store_str else True
            py_weekday = dv.weekday()
            js_index = (py_weekday + 1) % 7
            weekday_abbr = JS_WEEKDAYS[js_index]
            weekend = weekday_abbr in ('Sun', 'Lau')
            days.append({
                'iso': dv.date().isoformat(),
                'day': dv.day,
                'weekday': weekday_abbr,
                'week': str(week_val) if week_val is not None else '',
                'manager': manager_str,
                'store': store_str,
                'back': back_str,
                'closed': closed,
                'weekend': weekend,
                'campaign': str(campaign_val).strip() if campaign_val else '',
                'event': str(event_val).strip() if event_val else '',
            })

        # starfsmannalínur + deildafyrirsagnir í upprunalegri röð
        # (notum fyrirsagnir sem standa í blaðinu sjálfu, t.d. "HOME","TECH","VERSLUNARSTJÓRI",
        #  frekar en Staff-blaðið sem er ófullkomið fyrir Deild-dálkinn hjá mörgum)
        segments = []  # [{dept, rows:[(row, initials)]}]
        current_dept = None
        current_group = None
        for r in range(row_dags + 1, ws.max_row + 1):
            c_val = ws.cell(row=r, column=3).value
            if c_val is None or str(c_val).strip() == '':
                continue
            c_str = str(c_val).strip()
            if c_str in staff_initials:
                if current_group is None:
                    current_group = {'dept': current_dept or 'Annað', 'rows': []}
                    segments.append(current_group)
                current_group['rows'].append((r, c_str))
            elif c_str.lower().startswith('samtals'):
                continue  # eigin samtölur reiknaðar sjálf
            else:
                current_dept = c_str
                current_group = None  # næsta starfsmannalína byrjar nýjan hóp

        rows = []
        n_days = len(day_cols)
        overall_totals = [0] * n_days
        month_status_counts = {'shift': 0, 'summer': 0, 'parental': 0, 'unpaid': 0, 'sick': 0, 'other': 0}

        for seg in segments:
            dept = seg['dept']
            group = seg['rows']
            if not group:
                continue
            rows.append({'type': 'dept', 'dept': dept, 'initials': dept, 'name': dept, 'phone': '', 'values': [''] * n_days})
            dept_total = [0] * n_days
            for r, ini in group:
                vals = []
                for di, (col, dv) in enumerate(day_cols):
                    raw = ws.cell(row=r, column=col).value
                    v = str(raw).strip() if raw is not None else ''
                    vals.append(v)
                    kind = status_kind(v)
                    if kind == 'empty':
                        continue
                    if kind not in month_status_counts:
                        kind = 'other'
                    month_status_counts[kind] += 1
                    if kind == 'shift':
                        dept_total[di] += 1
                        overall_totals[di] += 1
                rows.append({
                    'type': 'employee', 'dept': dept,
                    'initials': ini,
                    'name': name_map.get(ini, ini),
                    'phone': phone_map.get(ini, ''),
                    'values': vals,
                })
            rows.append({'type': 'total', 'dept': dept, 'initials': '', 'name': f'Samtals {dept}', 'phone': '', 'values': [str(x) for x in dept_total]})

        rows.append({'type': 'total', 'dept': 'ALL', 'initials': '', 'name': 'Samtals í vinnu', 'phone': '', 'values': [str(x) for x in overall_totals]})

        closed_days = sum(1 for d in days if d['closed'])
        weekends = sum(1 for d in days if d['weekend'])
        avg_staff = round(sum(overall_totals) / len(overall_totals), 1) if overall_totals else 0

        months[f'{mname} {y}'] = {
            'locked': False,
            'days': days,
            'rows': rows,
            'totals': overall_totals,
            'summary': {
                'avgStaff': avg_staff,
                'maxStaff': max(overall_totals) if overall_totals else 0,
                'minStaff': min(overall_totals) if overall_totals else 0,
                'days': len(days),
                'closedDays': closed_days,
                'weekends': weekends,
                'statusCounts': month_status_counts,
            }
        }

    return months, skipped

if __name__ == '__main__':
    xlsx = sys.argv[1] if len(sys.argv) > 1 else 'Vaktaplan_Skeifan.xlsx'
    out = sys.argv[2] if len(sys.argv) > 2 else 'full-months.json'
    namemap_path = sys.argv[3] if len(sys.argv) > 3 else 'name_map.json'
    nm = json.load(open(namemap_path, encoding='utf-8'))
    months, skipped = build_full_months(xlsx, nm['names'], nm['phones'])
    json.dump(months, open(out, 'w', encoding='utf-8'), ensure_ascii=False, separators=(',', ':'))
    print(f'Skrifaði {len(months)} mánuði í {out}')
    if skipped:
        print(f'Slepptum: {skipped}')
