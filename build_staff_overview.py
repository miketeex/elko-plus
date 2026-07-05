#!/usr/bin/env python3
"""
Byggir staff_overview.json (tímabila-uppbygging) úr Vinnuyfirlit_Margir.xlsx.
Hvert innflutt Excel-skjal verður eitt "tímabil" (period) í úttakinu, svo hægt sé
að bera saman marga mánuði/ár þegar fleiri skjöl bætast við.

Aldur er reiknaður út frá kennitölu en kennitalan sjálf er ALDREI vistuð né birt.
"""
import openpyxl, re, json, datetime, sys

NAME_ALIASES = {
    'Nanna Katrín Snorradóttir': 'Nanna Kristín Snorradóttir',
    'Magnús Orri Sigmundsson': 'Magnús Orri Sigumundsson',
}

HEADER_RE = re.compile(r'^(.*?)\s*,,\s*kt\.\s*(\d{10})\s*,\s*(\d{2}\.\d{2}\.\d{4})\s*-\s*(\d{2}\.\d{2}\.\d{4})')
DAY_RE = re.compile(r'^(\d{2})\.(\d{2})\s*([a-záðéíóúýþæö]?)$', re.IGNORECASE)
FOOTNOTE_RE = re.compile(r'^\s*(\d+)\)\s*(.*)$')

def classify_action(code):
    c = (code or '').strip().upper().replace('Ó', 'O')
    if c in ('VINNA', 'AUKA'):
        return 'work'
    if c == 'VFL':
        return 'sick_stadfest'
    if c == 'VEB':
        return 'sick_barn'
    if c == 'VEIKOST':
        return 'sick_ostadfest'
    if c == 'LL':
        return 'unpaid'
    if c == 'SUM':
        return 'summer'
    if c == 'JFO':
        return 'funeral'
    if c == 'NAMSK':
        return 'training'
    if c == 'LOKAÐ':
        return 'closed'
    if c == 'LAUL':
        return 'unpaid'
    if c:
        return 'other'
    return 'empty'

DEPT_KEYWORDS = [
    ('KnowHow', 'Tækniþjónusta'),
    ('Verslunarstjóri', 'STJÓRI'),
    ('TECH', 'TECH'),
    ('HOME', 'HOME'),
    ('Þjónustudeild', 'ÞJÓNÓ'),
    ('Þf', 'ÞJÓNÓ'),
    ('Lager', 'LAGER'),
]
def normalize_dept(eining):
    if not eining:
        return ''
    e = str(eining)
    for kw, canon in DEPT_KEYWORDS:
        if kw in e:
            return canon
    return e.replace('Skeifan', '').strip()

def build_name_to_initials(staff_xlsx):
    wb = openpyxl.load_workbook(staff_xlsx, data_only=True)
    ws = wb['Staff']
    m = {}
    for r in range(2, ws.max_row + 1):
        ini = ws.cell(row=r, column=1).value
        full = ws.cell(row=r, column=2).value
        if ini and full:
            m[str(full).strip()] = str(ini).strip()
    return m

def kennitala_age(kt, reference_date):
    dd, mm, yy = int(kt[0:2]), int(kt[2:4]), int(kt[4:6])
    for century in (1900, 2000):
        try:
            byear = century + yy
            bdate = datetime.date(byear, mm, dd)
        except ValueError:
            continue
        age = reference_date.year - bdate.year - ((reference_date.month, reference_date.day) < (bdate.month, bdate.day))
        if 14 <= age <= 80:
            return age
    return None

def parse_time_range(s):
    if not s or '-' not in str(s):
        return None, None
    a, b = str(s).split('-', 1)
    def to_min(t):
        t = t.strip()
        if ':' not in t:
            return None
        h, m = t.split(':')
        return int(h) * 60 + int(m)
    return to_min(a), to_min(b)

def parse_footnotes(ws):
    last_row = ws.max_row
    for r in range(last_row, max(last_row - 4, 1), -1):
        val = ws.cell(row=r, column=2).value
        if val and ')' in str(val):
            notes = {}
            for line in str(val).split('\n'):
                m = FOOTNOTE_RE.match(line)
                if m:
                    notes[m.group(1)] = m.group(2).strip(' |')
            if notes:
                return notes
    return {}

def build_col_map(ws, header_row=3):
    col_map = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=c).value
        if v:
            col_map[str(v).strip()] = c
    return col_map

def parse_sheet(ws, year_hint_start, year_hint_end):
    rows = []
    footnotes = parse_footnotes(ws)
    cols = build_col_map(ws)
    c_dag = cols.get('Dag.', 2)
    c_eining = cols.get('Eining')
    c_adgerd = cols.get('Adgerd')
    c_skipulag = cols.get('Skipulag')
    c_vidvera = cols.get('Vidvera')
    c_ath = cols.get('Ath')  # ekki alltaf til staðar - fer eftir starfsmanni/tímabili
    r = 4
    max_r = ws.max_row
    while r <= max_r:
        dag_raw = ws.cell(row=r, column=c_dag).value
        if dag_raw:
            m = DAY_RE.match(str(dag_raw).strip())
            if m:
                dd, mm, wd = m.groups()
                month = int(mm)
                year = year_hint_start.year if month == year_hint_start.month else year_hint_end.year
                try:
                    date = datetime.date(year, month, int(dd))
                except ValueError:
                    r += 1
                    continue
                eining = ws.cell(row=r, column=c_eining).value if c_eining else None
                adgerd = ws.cell(row=r, column=c_adgerd).value if c_adgerd else None
                skipulag = ws.cell(row=r, column=c_skipulag).value if c_skipulag else None
                vidvera = ws.cell(row=r, column=c_vidvera).value if c_vidvera else None
                ath = ws.cell(row=r, column=c_ath).value if c_ath else None
                ath_str = str(ath).strip() if ath else ''
                ath_num_m = re.match(r'^(\d+)\)', ath_str)
                comment = footnotes.get(ath_num_m.group(1)) if ath_num_m else (ath_str or None)
                rows.append({
                    'date': date.isoformat(),
                    'eining': str(eining).strip() if eining else '',
                    'adgerd': str(adgerd).strip() if adgerd else '',
                    'skipulag': str(skipulag).strip() if skipulag else '',
                    'vidvera': str(vidvera).strip() if vidvera else '',
                    'comment': comment,
                })
        r += 1
    return rows

def build_period(vinnuyfirlit_xlsx, staff_xlsx, name_map, phone_map):
    name_to_initials = build_name_to_initials(staff_xlsx)
    wb = openpyxl.load_workbook(vinnuyfirlit_xlsx, data_only=True)

    by_kt = {}
    skipped_sheets = []
    period_start, period_end = None, None

    for sheetname in wb.sheetnames:
        ws = wb[sheetname]
        header = ws.cell(row=2, column=2).value
        if not header:
            skipped_sheets.append(sheetname)
            continue
        m = HEADER_RE.match(str(header))
        if not m:
            skipped_sheets.append(sheetname)
            continue
        fullname, kt, d1, d2 = m.groups()
        fullname = fullname.strip()
        lookup_name = NAME_ALIASES.get(fullname, fullname)
        initials = name_to_initials.get(lookup_name)
        if not initials:
            skipped_sheets.append(sheetname + ' (nafn fannst ekki: ' + fullname + ')')
            continue
        start = datetime.datetime.strptime(d1, '%d.%m.%Y').date()
        end = datetime.datetime.strptime(d2, '%d.%m.%Y').date()
        period_start, period_end = start, end

        rows = parse_sheet(ws, start, end)
        if not rows:
            continue

        entry = by_kt.setdefault(kt, {'fullname': fullname, 'initials': initials, 'rows_by_date': {}})
        for row in rows:
            entry['rows_by_date'][row['date']] = row

    employees = []
    for kt, entry in by_kt.items():
        ini = entry['initials']
        age = kennitala_age(kt, datetime.date.today())
        rows = sorted(entry['rows_by_date'].values(), key=lambda r: r['date'])

        worked_days = 0
        late_count = 0
        late_minutes_total = 0
        sick = {'stadfest': 0, 'barn': 0, 'ostadfest': 0}
        unpaid = 0
        training = 0
        summer = 0
        funeral = 0
        closed = 0
        other = 0
        depts = {}
        day_records = []

        for row in rows:
            kind = classify_action(row['adgerd'])
            dept = normalize_dept(row['eining'])
            if dept:
                depts[dept] = depts.get(dept, 0) + 1

            is_late = False
            late_min = 0
            if kind == 'work':
                worked_days += 1
                s_start, _ = parse_time_range(row['skipulag'])
                v_start, _ = parse_time_range(row['vidvera'])
                if s_start is not None and v_start is not None and v_start > s_start + 5:
                    is_late = True
                    late_min = v_start - s_start
                    late_count += 1
                    late_minutes_total += late_min
            elif kind == 'sick_stadfest':
                sick['stadfest'] += 1
            elif kind == 'sick_barn':
                sick['barn'] += 1
            elif kind == 'sick_ostadfest':
                sick['ostadfest'] += 1
            elif kind == 'unpaid':
                unpaid += 1
            elif kind == 'training':
                training += 1
            elif kind == 'summer':
                summer += 1
            elif kind == 'funeral':
                funeral += 1
            elif kind == 'closed':
                closed += 1
            elif kind == 'other':
                other += 1

            day_records.append({
                'date': row['date'], 'adgerd': row['adgerd'], 'dept': dept,
                'skipulag': row['skipulag'], 'vidvera': row['vidvera'],
                'late': is_late, 'lateMin': late_min, 'comment': row['comment'],
            })

        top_dept = max(depts.items(), key=lambda x: x[1])[0] if depts else ''
        total_sick = sick['stadfest'] + sick['barn'] + sick['ostadfest']
        day_records = [d for d in day_records if d['adgerd']]

        employees.append({
            'initials': ini,
            'name': name_map.get(ini, ini),
            'phone': phone_map.get(ini, ''),
            'dept': top_dept,
            'age': age,
            'workedDays': worked_days,
            'lateCount': late_count,
            'lateMinutesTotal': late_minutes_total,
            'sick': sick,
            'sickTotal': total_sick,
            'unpaid': unpaid,
            'training': training,
            'summer': summer,
            'funeral': funeral,
            'closed': closed,
            'other': other,
            'days': day_records,
        })

    employees.sort(key=lambda e: e['name'])
    period = {
        'label': f'{period_start.strftime("%d.%m.%Y")} - {period_end.strftime("%d.%m.%Y")}' if period_start else '',
        'start': period_start.isoformat() if period_start else None,
        'end': period_end.isoformat() if period_end else None,
        'importedAt': datetime.datetime.now().isoformat(timespec='minutes'),
        'employees': employees,
    }
    return period, skipped_sheets

def attach_seniority(period, first_seen_path):
    try:
        first_seen = json.load(open(first_seen_path, encoding='utf-8'))
    except FileNotFoundError:
        first_seen = {}
    MONTH_ORDER = ['Janúar','Febrúar','Mars','Apríl','Maí','Júní','Júlí','Ágúst','September','Október','Nóvember','Desember']
    today = datetime.date.today()
    def months_since(label):
        name, year = label.rsplit(' ', 1)
        midx = MONTH_ORDER.index(name)
        start = datetime.date(int(year), midx + 1, 1)
        return (today.year - start.year) * 12 + (today.month - start.month)
    for e in period['employees']:
        label = first_seen.get(e['initials'])
        e['startMonth'] = label
        e['seniorityMonths'] = months_since(label) if label else None

if __name__ == '__main__':
    vinnu = sys.argv[1] if len(sys.argv) > 1 else 'Vinnuyfirlit_Margir.xlsx'
    staff = sys.argv[2] if len(sys.argv) > 2 else 'Vaktaplan_Skeifan.xlsx'
    namemap_path = sys.argv[3] if len(sys.argv) > 3 else 'name_map.json'
    out = sys.argv[4] if len(sys.argv) > 4 else 'staff_overview.json'
    first_seen_path = sys.argv[5] if len(sys.argv) > 5 else 'first_seen.json'

    nm = json.load(open(namemap_path, encoding='utf-8'))
    period, skipped = build_period(vinnu, staff, nm['names'], nm['phones'])
    attach_seniority(period, first_seen_path)

    output = {'periods': [period]}
    json.dump(output, open(out, 'w', encoding='utf-8'), ensure_ascii=False, separators=(',', ':'))
    print(f'{len(period["employees"])} starfsmenn í tímabili {period["label"]} skrifaðir í {out}')
    print(f'Sleppt: {len(skipped)} blöðum')
    for s in skipped:
        print('  -', s)
