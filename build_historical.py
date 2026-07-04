#!/usr/bin/env python3
"""
Byggir historical.json úr Vaktaplan_Skeifan.xlsx fyrir ELKO Vaktaplan.
Keyrsla: python3 build_historical.py Vaktaplan_Skeifan.xlsx historical.json
"""
import sys, json, datetime
import openpyxl

MONTH_ORDER = ['Janúar','Febrúar','Mars','Apríl','Maí','Júní','Júlí','Ágúst','September','Október','Nóvember','Desember']
JS_WEEKDAYS = ['Sun','Mán','Þri','Mið','Fim','Fös','Lau']  # index = JS getDay() (Sunday=0)

def parse_month_sheet(name):
    for m in MONTH_ORDER:
        if name.startswith(m):
            try:
                year = int(name.split()[-1])
            except ValueError:
                return None, None
            return m, year
    return None, None

def find_label_row(ws, label, max_search_row=10, max_search_col=10):
    for r in range(1, max_search_row+1):
        for c in range(1, max_search_col+1):
            if ws.cell(row=r, column=c).value == label:
                return r, c
    return None, None

def status_kind(v):
    if v is None or str(v).strip() == '':
        return 'empty'
    s = str(v).strip().lower()
    if s == 'x':
        return 'unpaid'
    if s in ('frí', 'fri'):
        return 'summer'
    if 'f.orlof' in s:
        return 'parental'
    if s == 'v' or 'veik' in s:
        return 'sick'
    if any(ch.isdigit() for ch in s):
        return 'shift'
    return 'other'

def build_staff_map(wb):
    ws = wb['Staff']
    m = {}
    for r in range(2, ws.max_row + 1):
        sst = ws.cell(row=r, column=1).value
        dept = ws.cell(row=r, column=6).value
        if sst:
            m[str(sst).strip()] = str(dept).strip() if dept else ''
    return m

def build_historical(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    staff_map = build_staff_map(wb)
    staff_initials = set(staff_map.keys())

    sheets = []
    for s in wb.sheetnames:
        clean = s.replace('Í vinnslu - ', '')
        m, y = parse_month_sheet(clean)
        if m:
            sheets.append((s, m, y, MONTH_ORDER.index(m)))
    sheets.sort(key=lambda x: (x[1], x[2]))

    historical = []
    skipped = []

    for sheetname, mname, y, midx in sheets:
        ws = wb[sheetname]
        row_v, label_col = find_label_row(ws, 'Vaktstjóri')
        if not row_v:
            skipped.append(sheetname)
            continue
        row_dags = None
        for rr in range(row_v + 1, row_v + 4):
            if ws.cell(row=rr, column=label_col).value == 'Dags':
                row_dags = rr
                break
        if not row_dags:
            skipped.append(sheetname)
            continue
        row_store, _ = find_label_row(ws, 'Verslanir', max_search_row=row_dags + 4, max_search_col=label_col)
        col_start = label_col + 1
        max_col = ws.max_column

        # Collect employee rows: any row where column C (index label_col-3, i.e. col 3) holds a known staff initial
        # Column layout observed: col C = initials (col index 3), col D = name (col index 4)
        employee_rows = []
        for r in range(row_dags + 1, ws.max_row + 1):
            initials_val = ws.cell(row=r, column=3).value
            if initials_val is not None and str(initials_val).strip() in staff_initials:
                employee_rows.append((r, str(initials_val).strip()))

        col = col_start
        while col <= max_col:
            dv = ws.cell(row=row_dags, column=col).value
            if isinstance(dv, datetime.datetime) and dv.month == midx + 1 and dv.year == y:
                store_val = ws.cell(row=row_store, column=col).value if row_store else None
                store_str = str(store_val).strip() if store_val is not None else ''
                closed = 'lokað' in store_str.lower() if store_str else True

                status_counts = {'shift': 0, 'summer': 0, 'parental': 0, 'unpaid': 0, 'sick': 0, 'other': 0}
                dept_counts = {}
                for r, ini in employee_rows:
                    val = ws.cell(row=r, column=col).value
                    kind = status_kind(val)
                    if kind == 'empty':
                        continue
                    if kind not in status_counts:
                        kind = 'other'
                    status_counts[kind] += 1
                    if kind == 'shift':
                        dept = staff_map.get(ini, '')
                        if dept:
                            dept_counts[dept] = dept_counts.get(dept, 0) + 1

                iso = dv.date().isoformat()
                py_weekday = dv.weekday()  # Monday=0 ... Sunday=6
                js_index = (py_weekday + 1) % 7
                weekday_abbr = JS_WEEKDAYS[js_index]
                weekend = weekday_abbr in ('Sun', 'Lau')

                historical.append({
                    'date': iso,
                    'weekday': weekday_abbr,
                    'weekend': weekend,
                    'closed': closed,
                    'store': store_str if store_str else '',
                    'staff': status_counts['shift'],
                    'monthName': f'{mname} {y}',
                    'day': dv.day,
                    'deptCounts': dept_counts,
                    'statusCounts': status_counts,
                })
            col += 1

    historical.sort(key=lambda d: d['date'])
    return historical, skipped

if __name__ == '__main__':
    src = sys.argv[1] if len(sys.argv) > 1 else 'Vaktaplan_Skeifan.xlsx'
    dst = sys.argv[2] if len(sys.argv) > 2 else 'historical.json'
    data, skipped = build_historical(src)
    with open(dst, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, separators=(',', ':'))
    print(f'Skrifaði {len(data)} daga í {dst}')
    if skipped:
        print(f'Slepptum blöðum (ekkert Vaktstjóri/Dags snið fundið): {skipped}')
